# -*- coding: utf-8 -*-
from odoo import _, api, fields, models
from odoo.exceptions import UserError
import io

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    openpyxl = None


class ScannerCycleCount(models.Model):
    _name = "scanner.cycle.count"
    _description = "Sesión de Conteo Cíclico en Escáner Móvil"
    _order = "id desc"

    name = fields.Char(
        string="Código de Conteo",
        required=True,
        default=lambda self: _("Nuevo"),
        copy=False,
        readonly=True,
    )
    user_id = fields.Many2one(
        "res.users",
        string="Operador / Usuario",
        default=lambda self: self.env.user,
        required=True,
    )
    company_id = fields.Many2one(
        "res.company",
        string="Compañía",
        default=lambda self: self.env.company,
        required=True,
    )
    date = fields.Datetime(
        string="Fecha de Inicio",
        default=fields.Datetime.now,
        required=True,
    )
    date_end = fields.Datetime(
        string="Fecha Finalización",
    )
    state = fields.Selection(
        [
            ("draft", "Borrador"),
            ("in_progress", "En Proceso"),
            ("completed", "Finalizado"),
            ("cancelled", "Cancelado"),
        ],
        string="Estado",
        default="draft",
        required=True,
        index=True,
    )
    line_ids = fields.One2many(
        "scanner.cycle.count.line",
        "cycle_count_id",
        string="Líneas de Conteo",
    )

    total_lines = fields.Integer(
        string="Productos Distintos",
        compute="_compute_totals",
        store=True,
    )
    total_system_qty = fields.Float(
        string="Total Stock Inicial (Sistema)",
        compute="_compute_totals",
        store=True,
    )
    total_counted_qty = fields.Float(
        string="Total Stock Contado",
        compute="_compute_totals",
        store=True,
    )
    total_diff_qty = fields.Float(
        string="Diferencia Neta (Uds)",
        compute="_compute_totals",
        store=True,
    )
    total_diff_usd = fields.Float(
        string="Diferencia Monetaria ($ USD)",
        compute="_compute_totals",
        store=True,
    )

    @api.depends("line_ids", "line_ids.system_qty", "line_ids.counted_qty", "line_ids.difference_value_usd")
    def _compute_totals(self):
        for rec in self:
            lines = rec.line_ids
            rec.total_lines = len(lines)
            rec.total_system_qty = sum(lines.mapped("system_qty"))
            rec.total_counted_qty = sum(lines.mapped("counted_qty"))
            rec.total_diff_qty = rec.total_counted_qty - rec.total_system_qty
            rec.total_diff_usd = sum(lines.mapped("difference_value_usd"))

    @api.model_create_multi
    def create(self, vals_list):
        for vals in vals_list:
            if vals.get("name", _("Nuevo")) == _("Nuevo"):
                seq_date = fields.Date.today()
                vals["name"] = self.env["ir.sequence"].next_by_code(
                    "scanner.cycle.count", sequence_date=seq_date
                ) or f"CC/{fields.Date.today().year}/{self.search_count([]) + 1:05d}"
        return super().create(vals_list)

    def action_start(self):
        self.write({"state": "in_progress", "date": fields.Datetime.now()})

    def action_finish(self):
        self.write({"state": "completed", "date_end": fields.Datetime.now()})

    def action_cancel(self):
        self.write({"state": "cancelled"})

    def action_reset_to_draft(self):
        self.write({"state": "draft"})

    def export_excel_report_bytes(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(_("La librería 'openpyxl' no está disponible en el servidor."))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Conteo Cíclico"

        # Styles
        font_title = Font(name="Calibri", size=16, bold=True, color="1E293B")
        font_subtitle = Font(name="Calibri", size=10, italic=True, color="64748B")
        font_meta_label = Font(name="Calibri", size=10, bold=True, color="0F172A")
        font_meta_val = Font(name="Calibri", size=10, color="334155")
        
        fill_header = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        fill_summary = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        font_summary_title = Font(name="Calibri", size=11, bold=True, color="0F172A")
        font_summary_val = Font(name="Calibri", size=12, bold=True, color="0284C7")

        font_data = Font(name="Calibri", size=10, color="0F172A")
        font_diff_pos = Font(name="Calibri", size=10, bold=True, color="0284C7") # Surplus
        font_diff_neg = Font(name="Calibri", size=10, bold=True, color="DC2626") # Deficit
        font_diff_zero = Font(name="Calibri", size=10, color="16A34A")          # Exact match

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")
        align_right = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style="thin", color="CBD5E1"),
            right=Side(style="thin", color="CBD5E1"),
            top=Side(style="thin", color="CBD5E1"),
            bottom=Side(style="thin", color="CBD5E1"),
        )

        # 1. Title Block
        ws.merge_cells("A1:K1")
        ws["A1"] = "REPORTE DE CONTEO CÍCLICO DE INVENTARIO"
        ws["A1"].font = font_title
        ws["A1"].alignment = align_left

        ws.merge_cells("A2:K2")
        ws["A2"] = f"Generado desde Escáner Móvil Pro - {self.company_id.name}"
        ws["A2"].font = font_subtitle
        ws["A2"].alignment = align_left

        # 2. Metadata Block
        ws["A4"] = "Código Sesión:"
        ws["A4"].font = font_meta_label
        ws["B4"] = self.name
        ws["B4"].font = font_meta_val

        ws["D4"] = "Operador:"
        ws["D4"].font = font_meta_label
        ws["E4"] = self.user_id.name
        ws["E4"].font = font_meta_val

        ws["G4"] = "Fecha Inicio:"
        ws["G4"].font = font_meta_label
        ws["H4"] = self.date.strftime("%Y-%m-%d %H:%M:%S") if self.date else "N/A"
        ws["H4"].font = font_meta_val

        ws["A5"] = "Estado:"
        ws["A5"].font = font_meta_label
        ws["B5"] = dict(self._fields["state"].selection).get(self.state, self.state)
        ws["B5"].font = font_meta_val

        ws["D5"] = "Fecha Final:"
        ws["D5"].font = font_meta_label
        ws["E5"] = self.date_end.strftime("%Y-%m-%d %H:%M:%S") if self.date_end else "En Proceso"
        ws["E5"].font = font_meta_val

        # 3. Summary Cards Block
        ws.merge_cells("A7:B7")
        ws["A7"] = "PRODUCTOS DISTINTOS"
        ws["A7"].font = font_summary_title
        ws["A7"].fill = fill_summary
        ws["A7"].alignment = align_center

        ws.merge_cells("A8:B8")
        ws["A8"] = self.total_lines
        ws["A8"].font = font_summary_val
        ws["A8"].alignment = align_center

        ws.merge_cells("D7:E7")
        ws["D7"] = "STOCK INICIAL (SISTEMA)"
        ws["D7"].font = font_summary_title
        ws["D7"].fill = fill_summary
        ws["D7"].alignment = align_center

        ws.merge_cells("D8:E8")
        ws["D8"] = self.total_system_qty
        ws["D8"].font = font_summary_val
        ws["D8"].alignment = align_center

        ws.merge_cells("G7:H7")
        ws["G7"] = "STOCK CONTADO (CÍCLICO)"
        ws["G7"].font = font_summary_title
        ws["G7"].fill = fill_summary
        ws["G7"].alignment = align_center

        ws.merge_cells("G8:H8")
        ws["G8"] = self.total_counted_qty
        ws["G8"].font = font_summary_val
        ws["G8"].alignment = align_center

        ws.merge_cells("J7:K7")
        ws["J7"] = "DIFERENCIA MONETARIA ($ USD)"
        ws["J7"].font = font_summary_title
        ws["J7"].fill = fill_summary
        ws["J7"].alignment = align_center

        ws.merge_cells("J8:K8")
        ws["J8"] = self.total_diff_usd
        ws["J8"].font = font_summary_val
        ws["J8"].number_format = "$#,##0.00;($#,##0.00);$0.00"
        ws["J8"].alignment = align_center

        # 4. Table Headers
        headers = [
            "N°",
            "Código de Barras",
            "Referencia / SKU",
            "Descripción del Producto",
            "Ubicación",
            "Stock Inicial (Sistema)",
            "Stock Contado (Cíclico)",
            "Diferencia (Uds)",
            "Precio Unit ($ USD)",
            "Valor Contado ($ USD)",
            "Diferencia ($ USD)",
        ]

        row_idx = 10
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        # 5. Data Rows
        for i, line in enumerate(self.line_ids, 1):
            row_idx += 1
            ws.cell(row=row_idx, column=1, value=i).alignment = align_center
            ws.cell(row=row_idx, column=2, value=line.barcode or "N/A").alignment = align_center
            ws.cell(row=row_idx, column=3, value=line.default_code or "N/A").alignment = align_center
            ws.cell(row=row_idx, column=4, value=line.product_name or "").alignment = align_left
            ws.cell(row=row_idx, column=5, value=line.location_name or "Sin Ubicación").alignment = align_center

            c_sys = ws.cell(row=row_idx, column=6, value=line.system_qty)
            c_sys.alignment = align_right
            c_sys.number_format = "#,##0"

            c_cnt = ws.cell(row=row_idx, column=7, value=line.counted_qty)
            c_cnt.alignment = align_right
            c_cnt.number_format = "#,##0"

            c_diff = ws.cell(row=row_idx, column=8, value=line.difference_qty)
            c_diff.alignment = align_right
            c_diff.number_format = "+#,##0;-#,##0;0"

            if line.difference_qty > 0:
                c_diff.font = font_diff_pos
            elif line.difference_qty < 0:
                c_diff.font = font_diff_neg
            else:
                c_diff.font = font_diff_zero

            c_price = ws.cell(row=row_idx, column=9, value=line.unit_price_usd)
            c_price.alignment = align_right
            c_price.number_format = "$#,##0.00"

            c_val_cnt = ws.cell(row=row_idx, column=10, value=line.counted_value_usd)
            c_val_cnt.alignment = align_right
            c_val_cnt.number_format = "$#,##0.00"

            c_val_diff = ws.cell(row=row_idx, column=11, value=line.difference_value_usd)
            c_val_diff.alignment = align_right
            c_val_diff.number_format = "$#,##0.00;($#,##0.00);$0.00"
            if line.difference_value_usd > 0:
                c_val_diff.font = font_diff_pos
            elif line.difference_value_usd < 0:
                c_val_diff.font = font_diff_neg
            else:
                c_val_diff.font = font_diff_zero

            for col_idx in range(1, 12):
                ws.cell(row=row_idx, column=col_idx).border = thin_border
                if col_idx not in (8, 11):
                    ws.cell(row=row_idx, column=col_idx).font = font_data

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()


class ScannerCycleCountLine(models.Model):
    _name = "scanner.cycle.count.line"
    _description = "Línea de Conteo Cíclico en Escáner Móvil"
    _order = "id asc"

    cycle_count_id = fields.Many2one(
        "scanner.cycle.count",
        string="Conteo Cíclico",
        ondelete="cascade",
        required=True,
    )
    product_id = fields.Many2one(
        "product.product",
        string="Producto",
        required=True,
    )
    barcode = fields.Char(string="Código de Barras")
    default_code = fields.Char(string="Referencia / SKU")
    product_name = fields.Char(string="Nombre del Producto")
    location_name = fields.Char(string="Ubicación Almacén")

    system_qty = fields.Float(string="Stock Inicial (Sistema)", default=0.0)
    counted_qty = fields.Float(string="Stock Contado", default=1.0)
    difference_qty = fields.Float(
        string="Diferencia (Uds)",
        compute="_compute_line_amounts",
        store=True,
    )

    unit_price_usd = fields.Float(string="Precio Unit. ($ USD)", default=0.0)
    system_value_usd = fields.Float(
        string="Valor Sistema ($ USD)",
        compute="_compute_line_amounts",
        store=True,
    )
    counted_value_usd = fields.Float(
        string="Valor Contado ($ USD)",
        compute="_compute_line_amounts",
        store=True,
    )
    difference_value_usd = fields.Float(
        string="Diferencia ($ USD)",
        compute="_compute_line_amounts",
        store=True,
    )

    @api.depends("system_qty", "counted_qty", "unit_price_usd")
    def _compute_line_amounts(self):
        for line in self:
            line.difference_qty = line.counted_qty - line.system_qty
            line.system_value_usd = line.system_qty * line.unit_price_usd
            line.counted_value_usd = line.counted_qty * line.unit_price_usd
            line.difference_value_usd = line.difference_qty * line.unit_price_usd
